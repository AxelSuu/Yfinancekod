from pathlib import Path
import yfinance as yf
from openpyxl import load_workbook

class FinData():
    def __init__(self, tickers: list):
        self.xlsx = load_workbook(Path(__file__).parent / 'findata.xlsx')
        self.sheet = self.xlsx['tickerdata']
        self.tickers = tickers
    
    def load_ticker(self, row: int, ticker: str):
        t_par = yf.Ticker(ticker)
        t_info = t_par.info
        assign(self.sheet, row, 1, ticker)
        assign(self.sheet, row, 2, t_par.info.get('regularMarketPrice'))
        assign(self.sheet, row, 3, str_from_timestamp(t_par.info.get('exDividendDate')))
        assign(self.sheet, row, 4, t_par.info.get('dividendYield'))
        assign(self.sheet, row, 5, t_par.info.get('priceToBook'))
        assign(self.sheet, row, 6, t_par.info.get('payoutRatio'))
        (e_date_from, e_date_to) = get_earn_dates(t_par)
        assign(self.sheet, row, 7, e_date_from)
        assign(self.sheet, row, 8, e_date_to)
    
    
    def load_all_tickers(self):
        for row, ticker in enumerate(self.tickers, start=2):
            self.load_ticker(row, ticker)
    
    def save_fin_data(self, name_xls):
        self.xlsx.save(Path(__file__).parent / name_xls)

    def assign(sheet, row, col, val):
        if val is not None:
            sheet.cell(row=row, column=col).value = val

    def str_from_timestamp(datim):
        if (datim is None) or (datim==''):
            return ''
        dt=datetime.utcfromtimestamp(int(datim))
        return dt.strftime('%Y-%m-%d')

    def get_earn_dates(t_par):
        e_date_from = ''
        e_date_to = ''
        cal = t_par.calendar
        if cal is not None:
            v = cal.values[0]
        if v is not None:
            lenV = len(v)
            if lenV > 0:
                e_date_from = date_from_timestamp(v[0])
                if lenV > 1:
                    e_date_from = date_from_timestamp(v[1])
        return (e_date_from, e_date_to)
    
tickers_list = ['AAPL','AMZN','GOOGL','MSFT','PFE','TSLA']
f_data = FinData(tickers_list)
f_data.load_all_tickers()
f_data.save_fin_data('findata01.xlsx')